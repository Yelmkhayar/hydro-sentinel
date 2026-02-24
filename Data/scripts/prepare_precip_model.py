#!/usr/bin/env python3
"""
Prepare and fill the multi-station precipitation template from a single model input CSV.

Usage example:
python prepare_precip_template.py \
  --input pcp/stations_timeseries_arome_2002.csv \
  --template template_multi_station_precip_mm_1802.xlsx \
  --model AROME \
  --outdir output
"""

from __future__ import annotations

import argparse
import json
import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl
import pandas as pd

REQUIRED_COLUMNS = ["time", "echance", "station_id", "name", "station_name", "rr"]
TIMESTAMP_FMT = "%Y-%m-%dT%H:%M:%S"


@dataclass
class ValidationResult:
    warnings: List[str]
    errors: List[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Prepare precipitation template for BD ingestion (single input model per run)."
    )
    parser.add_argument("--input", required=True, help="Input CSV path (AROME or ECMWF).")
    parser.add_argument("--template", required=True, help="Excel template path.")
    parser.add_argument("--model", required=True, help="Model label for traceability (e.g., AROME, ECMWF).")
    parser.add_argument("--outdir", default="outputs/runs", help="Output directory.")
    parser.add_argument("--sheet-data", default="Données", help="Template data sheet name.")
    parser.add_argument("--sheet-stations", default="Stations", help="Template station sheet name.")
    parser.add_argument(
        "--data-start-row",
        type=int,
        default=4,
        help="First data row in sheet-data (default: 4).",
    )
    parser.add_argument(
        "--fill-missing",
        type=float,
        default=0.0,
        help="Fill value for missing (time, station) after pivot (default: 0.0).",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail on warnings usually considered recoverable.",
    )
    return parser.parse_args()


def setup_logging(log_path: Path) -> logging.Logger:
    logger = logging.getLogger("prepare_precip_template")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    sh = logging.StreamHandler()
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


def read_input_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";")
    return df


def validate_and_clean_input(df: pd.DataFrame) -> Tuple[pd.DataFrame, ValidationResult, Dict[str, object]]:
    warnings: List[str] = []
    errors: List[str] = []

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing_cols:
        errors.append(f"Missing required columns: {missing_cols}")
        return df, ValidationResult(warnings, errors), {}

    raw_rows = len(df)

    clean = df.copy()
    clean["time"] = pd.to_datetime(clean["time"], errors="coerce")
    invalid_time = int(clean["time"].isna().sum())
    if invalid_time > 0:
        warnings.append(f"Invalid timestamps converted to NaT: {invalid_time}")
        clean = clean.dropna(subset=["time"])

    clean["station_id"] = pd.to_numeric(clean["station_id"], errors="coerce")
    invalid_station = int(clean["station_id"].isna().sum())
    if invalid_station > 0:
        warnings.append(f"Invalid station_id converted to NaN and dropped: {invalid_station}")
        clean = clean.dropna(subset=["station_id"])
    clean["station_id"] = clean["station_id"].astype(int)

    clean["rr"] = pd.to_numeric(clean["rr"], errors="coerce")
    invalid_rr = int(clean["rr"].isna().sum())
    if invalid_rr > 0:
        warnings.append(f"Non-numeric rr converted to NaN: {invalid_rr}")

    negative_rr = int((clean["rr"] < 0).sum(skipna=True))
    if negative_rr > 0:
        warnings.append(f"Negative rr values detected and kept as-is: {negative_rr}")

    dup_mask = clean.duplicated(subset=["time", "station_id"], keep="last")
    duplicate_rows_removed = int(dup_mask.sum())
    if duplicate_rows_removed > 0:
        warnings.append(
            f"Duplicate (time, station_id) rows removed (keep=last): {duplicate_rows_removed}"
        )
        clean = clean.loc[~dup_mask].copy()

    clean = clean.sort_values(["time", "station_id"]).reset_index(drop=True)

    stats = {
        "raw_rows": raw_rows,
        "rows_after_cleaning": int(len(clean)),
        "invalid_time_rows": invalid_time,
        "invalid_station_rows": invalid_station,
        "invalid_rr_rows": invalid_rr,
        "negative_rr_rows": negative_rr,
        "duplicate_rows_removed": duplicate_rows_removed,
        "input_station_count": int(clean["station_id"].nunique()) if len(clean) else 0,
        "input_time_count": int(clean["time"].nunique()) if len(clean) else 0,
        "input_time_min": clean["time"].min().isoformat() if len(clean) else None,
        "input_time_max": clean["time"].max().isoformat() if len(clean) else None,
    }

    if len(clean) == 0:
        errors.append("No valid rows remain after cleaning.")

    return clean, ValidationResult(warnings, errors), stats


def _extract_numeric_codes_from_row(ws, row_idx: int, start_col: int = 2) -> Dict[int, int]:
    code_to_col: Dict[int, int] = {}
    for col in range(start_col, ws.max_column + 1):
        value = ws.cell(row=row_idx, column=col).value
        if isinstance(value, (int, float)):
            code_to_col[int(value)] = col
    return code_to_col


def _extract_station_catalog(ws, code_col: int = 1, name_col: int = 2, start_row: int = 2) -> Dict[int, str]:
    catalog: Dict[int, str] = {}
    for row in range(start_row, ws.max_row + 1):
        code = ws.cell(row=row, column=code_col).value
        if isinstance(code, (int, float)):
            name = ws.cell(row=row, column=name_col).value
            catalog[int(code)] = str(name) if name is not None else ""
    return catalog


def build_station_order(
    existing_order: Sequence[int],
    stations_sheet_codes: Sequence[int],
    input_codes: Sequence[int],
) -> Tuple[List[int], List[int]]:
    ordered = list(existing_order)
    existing_set = set(existing_order)

    to_add = sorted((set(stations_sheet_codes) | set(input_codes)) - existing_set)
    ordered.extend(to_add)
    return ordered, to_add


def compute_time_gaps(times: pd.Series) -> List[str]:
    if times.empty:
        return []

    uniq = pd.Series(sorted(times.unique()))
    deltas = uniq.diff().dropna()
    gaps = deltas[deltas != pd.Timedelta(hours=1)]
    gap_messages: List[str] = []
    if gaps.empty:
        return gap_messages

    for idx, delta in gaps.items():
        t_prev = uniq.iloc[idx - 1]
        t_curr = uniq.iloc[idx]
        gap_messages.append(
            f"gap={delta} between {pd.Timestamp(t_prev).isoformat()} and {pd.Timestamp(t_curr).isoformat()}"
        )
    return gap_messages


def dataframe_to_matrix(
    clean: pd.DataFrame,
    target_station_order: Sequence[int],
    fill_missing: float,
) -> Tuple[pd.DataFrame, Dict[str, object]]:
    pivot = clean.pivot_table(index="time", columns="station_id", values="rr", aggfunc="last")
    pivot = pivot.sort_index()

    pre_fill_missing = int(pivot.reindex(columns=target_station_order).isna().sum().sum())
    matrix = pivot.reindex(columns=target_station_order).fillna(fill_missing)

    flat_values = matrix.to_numpy().ravel()
    rr_series = pd.Series(flat_values)

    stats = {
        "output_time_count": int(matrix.shape[0]),
        "output_station_count": int(matrix.shape[1]),
        "pre_fill_missing_cells": pre_fill_missing,
        "filled_missing_value": fill_missing,
        "rr_min": float(rr_series.min()) if len(rr_series) else None,
        "rr_max": float(rr_series.max()) if len(rr_series) else None,
        "rr_mean": float(rr_series.mean()) if len(rr_series) else None,
        "rr_p95": float(rr_series.quantile(0.95)) if len(rr_series) else None,
        "rr_p99": float(rr_series.quantile(0.99)) if len(rr_series) else None,
        "rr_zero_pct": float((rr_series == 0).mean() * 100) if len(rr_series) else None,
        "rr_non_zero_pct": float((rr_series > 0).mean() * 100) if len(rr_series) else None,
    }
    return matrix, stats


def clear_data_region(ws, start_row: int, max_col: int) -> None:
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col, value=None)


def write_template(
    wb: openpyxl.Workbook,
    data_sheet: str,
    matrix: pd.DataFrame,
    station_order: Sequence[int],
    data_start_row: int,
) -> Dict[str, int]:
    ws = wb[data_sheet]

    # Ensure headers
    ws.cell(row=3, column=1, value="timestamp")
    for idx, station_code in enumerate(station_order, start=2):
        ws.cell(row=3, column=idx, value=int(station_code))

    max_col = max(1, len(station_order) + 1)
    clear_data_region(ws, start_row=data_start_row, max_col=max_col)

    row = data_start_row
    for ts, values in matrix.iterrows():
        ws.cell(row=row, column=1, value=pd.Timestamp(ts).strftime(TIMESTAMP_FMT))
        for col_idx, val in enumerate(values.tolist(), start=2):
            ws.cell(row=row, column=col_idx, value=float(val))
        row += 1

    return {
        "written_rows": int(matrix.shape[0]),
        "written_station_columns": int(matrix.shape[1]),
        "first_output_row": data_start_row,
        "last_output_row": row - 1,
    }


def format_report_txt(report: Dict[str, object]) -> str:
    lines: List[str] = []
    lines.append("Transformation Report - Multi Station Precipitation")
    lines.append(f"run_id: {report['run_id']}")
    lines.append(f"model: {report['model']}")
    lines.append(f"input_file: {report['input_file']}")
    lines.append(f"template_file: {report['template_file']}")
    lines.append(f"output_file: {report['output_file']}")
    lines.append("")

    lines.append("Input Summary")
    input_stats = report["input_stats"]
    for key in sorted(input_stats.keys()):
        lines.append(f"- {key}: {input_stats[key]}")
    lines.append("")

    lines.append("Station Summary")
    station_stats = report["station_stats"]
    for key in sorted(station_stats.keys()):
        lines.append(f"- {key}: {station_stats[key]}")
    lines.append("")

    lines.append("Time Quality")
    tq = report["time_quality"]
    lines.append(f"- detected_gaps_count: {tq['detected_gaps_count']}")
    if tq["gaps"]:
        for g in tq["gaps"]:
            lines.append(f"  * {g}")
    lines.append("")

    lines.append("Output Summary")
    out = report["output_stats"]
    for key in sorted(out.keys()):
        lines.append(f"- {key}: {out[key]}")
    lines.append("")

    lines.append("Warnings")
    warnings = report["warnings"]
    if warnings:
        for w in warnings:
            lines.append(f"- {w}")
    else:
        lines.append("- none")

    return "\n".join(lines) + "\n"


def main() -> int:
    args = parse_args()

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    base_stem = f"precip_{args.model.lower()}_{run_id}"
    output_xlsx = outdir / f"template_multi_station_precip_mm_{args.model.lower()}_{run_id}.xlsx"
    report_json = outdir / f"{base_stem}_report.json"
    report_txt = outdir / f"{base_stem}_report.txt"
    log_path = outdir / f"{base_stem}.log"

    logger = setup_logging(log_path)

    input_path = Path(args.input)
    template_path = Path(args.template)

    if not input_path.exists():
        logger.error("Input file not found: %s", input_path)
        return 2
    if not template_path.exists():
        logger.error("Template file not found: %s", template_path)
        return 2

    logger.info("Reading input CSV: %s", input_path)
    input_df = read_input_csv(input_path)
    clean_df, validation, input_stats = validate_and_clean_input(input_df)

    if validation.errors:
        for err in validation.errors:
            logger.error(err)
        return 1

    logger.info("Loading template workbook: %s", template_path)
    wb = openpyxl.load_workbook(template_path)

    if args.sheet_data not in wb.sheetnames:
        logger.error("Data sheet not found: %s", args.sheet_data)
        return 1
    if args.sheet_stations not in wb.sheetnames:
        logger.error("Station sheet not found: %s", args.sheet_stations)
        return 1

    ws_data = wb[args.sheet_data]
    ws_st = wb[args.sheet_stations]

    existing_code_to_col = _extract_numeric_codes_from_row(ws_data, row_idx=3, start_col=2)
    existing_order = [code for code, _col in sorted(existing_code_to_col.items(), key=lambda kv: kv[1])]

    station_catalog = _extract_station_catalog(ws_st, code_col=1, name_col=2, start_row=2)
    station_sheet_codes = sorted(station_catalog.keys())

    input_codes = sorted(clean_df["station_id"].unique().tolist())

    target_station_order, added_codes = build_station_order(
        existing_order=existing_order,
        stations_sheet_codes=station_sheet_codes,
        input_codes=input_codes,
    )

    warnings = list(validation.warnings)
    if added_codes:
        warnings.append(
            "Station columns missing in template data sheet were added automatically: "
            + ", ".join(map(str, added_codes))
        )

    station_in_input_not_in_stations_sheet = sorted(set(input_codes) - set(station_sheet_codes))
    if station_in_input_not_in_stations_sheet:
        warnings.append(
            "Stations present in input but not in template Stations sheet: "
            + ", ".join(map(str, station_in_input_not_in_stations_sheet))
        )

    station_in_stations_sheet_not_in_input = sorted(set(station_sheet_codes) - set(input_codes))
    if station_in_stations_sheet_not_in_input:
        warnings.append(
            "Stations present in template Stations sheet but absent in input (filled with --fill-missing): "
            + ", ".join(map(str, station_in_stations_sheet_not_in_input))
        )

    time_gaps = compute_time_gaps(clean_df["time"])
    if time_gaps:
        warnings.append(f"Detected non-hourly time gaps: {len(time_gaps)}")

    matrix, matrix_stats = dataframe_to_matrix(
        clean=clean_df,
        target_station_order=target_station_order,
        fill_missing=args.fill_missing,
    )

    write_stats = write_template(
        wb=wb,
        data_sheet=args.sheet_data,
        matrix=matrix,
        station_order=target_station_order,
        data_start_row=args.data_start_row,
    )

    wb.save(output_xlsx)
    logger.info("Output workbook written: %s", output_xlsx)

    if args.strict and warnings:
        logger.error("Strict mode enabled and warnings were detected.")
        for w in warnings:
            logger.error("WARNING: %s", w)
        return 1

    for w in warnings:
        logger.warning(w)

    report = {
        "run_id": run_id,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "model": args.model,
        "input_file": str(input_path.resolve()),
        "template_file": str(template_path.resolve()),
        "output_file": str(output_xlsx.resolve()),
        "log_file": str(log_path.resolve()),
        "input_stats": input_stats,
        "station_stats": {
            "template_data_station_count_before": len(existing_order),
            "template_stations_sheet_count": len(station_sheet_codes),
            "input_station_count": len(input_codes),
            "target_station_count_after": len(target_station_order),
            "added_station_columns_in_data_sheet": added_codes,
            "stations_in_input_not_in_stations_sheet": station_in_input_not_in_stations_sheet,
            "stations_in_stations_sheet_not_in_input": station_in_stations_sheet_not_in_input,
        },
        "time_quality": {
            "detected_gaps_count": len(time_gaps),
            "gaps": time_gaps,
        },
        "output_stats": {
            **matrix_stats,
            **write_stats,
        },
        "warnings": warnings,
        "errors": validation.errors,
    }

    with report_json.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    with report_txt.open("w", encoding="utf-8") as f:
        f.write(format_report_txt(report))

    logger.info("Report JSON written: %s", report_json)
    logger.info("Report TXT written: %s", report_txt)
    logger.info("Done.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

