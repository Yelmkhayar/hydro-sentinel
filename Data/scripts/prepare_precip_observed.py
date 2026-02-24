#!/usr/bin/env python3
"""
Transform observed precipitation DataTable into the multi-station template format.
Supports:
- Excel DataTable (row1 codes + row2 names + data from row3)
- CSV DataTable (datetime column + station name columns)
"""

from __future__ import annotations

import argparse
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl
import pandas as pd

TIMESTAMP_FMT = "%Y-%m-%dT%H:%M:%S"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().replace("_", " ").split())


def parse_fill_value(raw: str):
    low = raw.strip().lower()
    if low in {"nan", "none", "null", ""}:
        return None
    return float(raw)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Prepare observed precipitation for BD ingestion template.")
    p.add_argument("--input", required=True, help="Input DataTable (.xlsx or .csv).")
    p.add_argument("--template", required=True, help="Template Excel path.")
    p.add_argument("--outdir", default="outputs/runs", help="Output folder.")
    p.add_argument("--sheet-data", default="Données", help="Template data sheet.")
    p.add_argument("--sheet-stations", default="Stations", help="Template stations sheet.")
    p.add_argument("--input-sheet", default="DataTable", help="Input sheet when input is .xlsx.")
    p.add_argument("--data-start-row", type=int, default=4, help="First data row in template data sheet.")
    p.add_argument(
        "--fill-missing",
        default="nan",
        help="Fill value for missing station/time cells (default: nan). Use 0 for numeric fill.",
    )
    p.add_argument("--strict", action="store_true", help="Fail if any warning is detected.")
    return p.parse_args()


def setup_logger(path: Path) -> logging.Logger:
    logger = logging.getLogger("prepare_precip_observed")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    sh = logging.StreamHandler()
    sh.setFormatter(fmt)
    sh.setLevel(logging.INFO)
    logger.addHandler(sh)

    fh = logging.FileHandler(path, encoding="utf-8")
    fh.setFormatter(fmt)
    fh.setLevel(logging.INFO)
    logger.addHandler(fh)
    return logger


def extract_template_stations(ws_st) -> Tuple[List[int], Dict[int, str], Dict[str, int]]:
    codes: List[int] = []
    code_to_label: Dict[int, str] = {}
    name_to_code: Dict[str, int] = {}

    for row in range(2, ws_st.max_row + 1):
        code = ws_st.cell(row=row, column=1).value
        if not isinstance(code, (int, float)):
            continue
        code = int(code)
        nom = ws_st.cell(row=row, column=2).value
        var_name = ws_st.cell(row=row, column=4).value

        codes.append(code)
        code_to_label[code] = str(var_name) if var_name else str(nom) if nom else str(code)

        variants = {
            normalize_text(var_name),
            normalize_text(nom),
            normalize_text(f"{nom}_Pluie 1hr (mm)" if nom else None),
            normalize_text(f"{nom} Pluie 1hr (mm)" if nom else None),
        }
        for v in variants:
            if not v:
                continue
            if v not in name_to_code:
                name_to_code[v] = code

    return sorted(set(codes)), code_to_label, name_to_code


def extract_template_data_codes(ws_data) -> List[int]:
    out: List[int] = []
    for col in range(2, ws_data.max_column + 1):
        v = ws_data.cell(row=3, column=col).value
        if isinstance(v, (int, float)):
            out.append(int(v))
    return out


def parse_excel_datatable(input_path: Path, input_sheet: str, warnings: List[str]) -> Tuple[pd.DataFrame, Dict[str, int]]:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if input_sheet not in wb.sheetnames:
        raise ValueError(f"Input sheet not found: {input_sheet}")
    ws = wb[input_sheet]

    first = ws.cell(row=1, column=1).value
    has_code_row = normalize_text(first) == "code"

    records: List[dict] = []
    recognized_cols = 0

    if has_code_row:
        # row 1 -> codes, row 2 -> labels, row >=3 -> data
        col_meta: List[Tuple[int, Optional[int], str]] = []
        for col in range(2, ws.max_column + 1):
            code_v = ws.cell(row=1, column=col).value
            lbl_v = ws.cell(row=2, column=col).value
            code = int(code_v) if isinstance(code_v, (int, float)) else None
            col_meta.append((col, code, str(lbl_v) if lbl_v is not None else ""))
            if code is not None:
                recognized_cols += 1

        for row in range(3, ws.max_row + 1):
            t = ws.cell(row=row, column=1).value
            if t is None:
                continue
            for col, code, lbl in col_meta:
                if code is None:
                    continue
                v = ws.cell(row=row, column=col).value
                records.append({"time": t, "station_id": code, "rr": v, "source_col": lbl})
    else:
        # header row with names in row1
        dt_header = ws.cell(row=1, column=1).value
        col_meta = []
        for col in range(2, ws.max_column + 1):
            lbl_v = ws.cell(row=1, column=col).value
            col_meta.append((col, str(lbl_v) if lbl_v is not None else ""))
            if lbl_v is not None:
                recognized_cols += 1

        warnings.append("Excel DataTable has no code row; station mapping will rely on station names.")

        for row in range(2, ws.max_row + 1):
            t = ws.cell(row=row, column=1).value
            if t is None:
                continue
            for col, lbl in col_meta:
                v = ws.cell(row=row, column=col).value
                records.append({"time": t, "station_id": None, "rr": v, "source_col": lbl})

    df = pd.DataFrame.from_records(records)
    stats = {
        "input_rows_raw": int(len(df)),
        "input_station_cols_raw": int(recognized_cols),
    }
    return df, stats


def parse_csv_datatable(input_path: Path) -> Tuple[pd.DataFrame, Dict[str, int]]:
    raw = pd.read_csv(input_path, sep=";", dtype=str)
    if raw.shape[1] < 2:
        raise ValueError("CSV must contain datetime column + at least one station column.")

    datetime_col = raw.columns[0]
    station_cols = list(raw.columns[1:])

    melted = raw.melt(id_vars=[datetime_col], value_vars=station_cols, var_name="source_col", value_name="rr")
    melted = melted.rename(columns={datetime_col: "time"})
    melted["station_id"] = None

    stats = {
        "input_rows_raw": int(len(melted)),
        "input_station_cols_raw": int(len(station_cols)),
    }
    return melted[["time", "station_id", "rr", "source_col"]], stats


def map_stations(df: pd.DataFrame, name_to_code: Dict[str, int], warnings: List[str]) -> Tuple[pd.DataFrame, Dict[str, object]]:
    out = df.copy()

    # If station_id already available, keep as primary mapping.
    from_code = out["station_id"].notna().sum()

    unresolved_before = out["station_id"].isna()
    if unresolved_before.any():
        mapped = out.loc[unresolved_before, "source_col"].map(lambda x: name_to_code.get(normalize_text(x)))
        out.loc[unresolved_before, "station_id"] = mapped

    unresolved_after = out["station_id"].isna()
    unresolved_names = sorted(out.loc[unresolved_after, "source_col"].dropna().unique().tolist())

    if unresolved_names:
        warnings.append(
            "Input columns not mapped to any station code (ignored): " + ", ".join(unresolved_names)
        )

    out = out.loc[~unresolved_after].copy()
    out["station_id"] = out["station_id"].astype(int)

    stats = {
        "rows_with_code_from_input": int(from_code),
        "rows_mapped_by_name": int((~unresolved_before & unresolved_after).sum()) if False else int(
            (df["station_id"].isna().sum() - unresolved_after.sum())
        ),
        "rows_unmapped_dropped": int(unresolved_after.sum()),
        "unmapped_input_columns": unresolved_names,
    }
    return out, stats


def clean_values(df: pd.DataFrame, warnings: List[str]) -> Tuple[pd.DataFrame, Dict[str, object]]:
    clean = df.copy()

    clean["time"] = pd.to_datetime(clean["time"], errors="coerce", dayfirst=True)
    invalid_time = int(clean["time"].isna().sum())
    if invalid_time:
        warnings.append(f"Invalid timestamps dropped: {invalid_time}")
        clean = clean.dropna(subset=["time"])

    clean["rr"] = clean["rr"].astype(str).str.replace(",", ".", regex=False)
    clean["rr"] = pd.to_numeric(clean["rr"], errors="coerce")
    invalid_rr = int(clean["rr"].isna().sum())
    if invalid_rr:
        warnings.append(f"Non-numeric rr converted to NaN: {invalid_rr}")

    neg_rr = int((clean["rr"] < 0).sum(skipna=True))
    if neg_rr:
        warnings.append(f"Negative rr values found: {neg_rr}")

    dup = clean.duplicated(subset=["time", "station_id"], keep="last")
    dup_count = int(dup.sum())
    if dup_count:
        warnings.append(f"Duplicate (time, station_id) dropped (keep=last): {dup_count}")
        clean = clean.loc[~dup]

    clean = clean.sort_values(["time", "station_id"]).reset_index(drop=True)

    stats = {
        "rows_after_cleaning": int(len(clean)),
        "invalid_time_rows": invalid_time,
        "invalid_rr_rows": invalid_rr,
        "negative_rr_rows": neg_rr,
        "duplicate_rows_removed": dup_count,
        "time_min": clean["time"].min().isoformat() if len(clean) else None,
        "time_max": clean["time"].max().isoformat() if len(clean) else None,
        "station_count_detected": int(clean["station_id"].nunique()) if len(clean) else 0,
        "time_count_detected": int(clean["time"].nunique()) if len(clean) else 0,
    }
    return clean, stats


def compute_time_gaps(times: pd.Series) -> List[str]:
    if times.empty:
        return []
    uniq = pd.Series(sorted(times.unique()))
    diffs = uniq.diff().dropna()
    gaps = diffs[diffs != pd.Timedelta(hours=1)]
    out: List[str] = []
    for idx, delta in gaps.items():
        out.append(f"gap={delta} between {uniq.iloc[idx-1]} and {uniq.iloc[idx]}")
    return out


def build_matrix(clean: pd.DataFrame, target_codes: Sequence[int], fill_missing) -> Tuple[pd.DataFrame, Dict[str, object]]:
    pivot = clean.pivot_table(index="time", columns="station_id", values="rr", aggfunc="last").sort_index()
    matrix = pivot.reindex(columns=target_codes)

    pre_fill_missing = int(matrix.isna().sum().sum())
    if fill_missing is not None:
        matrix = matrix.fillna(fill_missing)

    flat = pd.Series(matrix.to_numpy().ravel())
    stats = {
        "output_rows": int(matrix.shape[0]),
        "output_station_columns": int(matrix.shape[1]),
        "pre_fill_missing_cells": pre_fill_missing,
        "fill_missing_value": fill_missing,
        "rr_min": float(flat.min()) if len(flat.dropna()) else None,
        "rr_max": float(flat.max()) if len(flat.dropna()) else None,
        "rr_mean": float(flat.mean()) if len(flat.dropna()) else None,
        "rr_p95": float(flat.quantile(0.95)) if len(flat.dropna()) else None,
        "rr_p99": float(flat.quantile(0.99)) if len(flat.dropna()) else None,
    }
    return matrix, stats


def clear_data_region(ws, start_row: int, max_col: int) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c, value=None)


def write_output_template(
    wb,
    sheet_data: str,
    matrix: pd.DataFrame,
    target_codes: Sequence[int],
    data_start_row: int,
) -> Dict[str, int]:
    ws = wb[sheet_data]
    ws.cell(row=3, column=1, value="timestamp")
    for idx, code in enumerate(target_codes, start=2):
        ws.cell(row=3, column=idx, value=int(code))

    max_col = len(target_codes) + 1
    clear_data_region(ws, data_start_row, max_col)

    row = data_start_row
    for ts, values in matrix.iterrows():
        ws.cell(row=row, column=1, value=pd.Timestamp(ts).strftime(TIMESTAMP_FMT))
        for col, v in enumerate(values.tolist(), start=2):
            if pd.isna(v):
                ws.cell(row=row, column=col, value=None)
            else:
                ws.cell(row=row, column=col, value=float(v))
        row += 1

    return {
        "written_rows": int(matrix.shape[0]),
        "written_station_columns": int(matrix.shape[1]),
        "first_output_row": int(data_start_row),
        "last_output_row": int(row - 1),
    }


def report_txt(report: Dict[str, object]) -> str:
    lines = [
        "Transformation Report - Observed Precipitation DataTable",
        f"run_id: {report['run_id']}",
        f"input_file: {report['input_file']}",
        f"template_file: {report['template_file']}",
        f"output_file: {report['output_file']}",
        "",
        "Input Stats",
    ]
    for k, v in sorted(report["input_stats"].items()):
        lines.append(f"- {k}: {v}")

    lines.extend(["", "Mapping Stats"])
    for k, v in sorted(report["mapping_stats"].items()):
        lines.append(f"- {k}: {v}")

    lines.extend(["", "Cleaning Stats"])
    for k, v in sorted(report["cleaning_stats"].items()):
        lines.append(f"- {k}: {v}")

    lines.extend(["", "Output Stats"])
    for k, v in sorted(report["output_stats"].items()):
        lines.append(f"- {k}: {v}")

    lines.extend(["", "Time Quality"])
    lines.append(f"- gap_count: {report['time_quality']['gap_count']}")
    for gap in report["time_quality"]["gaps"]:
        lines.append(f"  * {gap}")

    lines.extend(["", "Warnings"])
    if report["warnings"]:
        for w in report["warnings"]:
            lines.append(f"- {w}")
    else:
        lines.append("- none")

    return "\n".join(lines) + "\n"


def main() -> int:
    args = parse_args()
    fill_missing = parse_fill_value(args.fill_missing)
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    input_path = Path(args.input)
    template_path = Path(args.template)
    if not input_path.exists():
        print(f"ERROR input not found: {input_path}")
        return 2
    if not template_path.exists():
        print(f"ERROR template not found: {template_path}")
        return 2

    base = f"precip_observed_{run_id}"
    log_path = outdir / f"{base}.log"
    out_xlsx = outdir / f"template_multi_station_precip_mm_observed_{run_id}.xlsx"
    out_json = outdir / f"{base}_report.json"
    out_txt = outdir / f"{base}_report.txt"

    logger = setup_logger(log_path)
    warnings: List[str] = []

    logger.info("Loading template: %s", template_path)
    wb_tpl = openpyxl.load_workbook(template_path)
    if args.sheet_data not in wb_tpl.sheetnames:
        logger.error("Template data sheet not found: %s", args.sheet_data)
        return 1
    if args.sheet_stations not in wb_tpl.sheetnames:
        logger.error("Template stations sheet not found: %s", args.sheet_stations)
        return 1

    ws_data = wb_tpl[args.sheet_data]
    ws_st = wb_tpl[args.sheet_stations]

    stations_sheet_codes, code_to_label, name_to_code = extract_template_stations(ws_st)
    data_sheet_codes = extract_template_data_codes(ws_data)

    logger.info("Reading input DataTable: %s", input_path)
    if input_path.suffix.lower() == ".xlsx":
        raw_df, input_stats = parse_excel_datatable(input_path, args.input_sheet, warnings)
    elif input_path.suffix.lower() == ".csv":
        raw_df, input_stats = parse_csv_datatable(input_path)
    else:
        logger.error("Unsupported input format: %s", input_path.suffix)
        return 1

    mapped_df, mapping_stats = map_stations(raw_df, name_to_code, warnings)
    clean_df, cleaning_stats = clean_values(mapped_df, warnings)
    if clean_df.empty:
        logger.error("No valid data after mapping/cleaning.")
        return 1

    target_codes = sorted(set(stations_sheet_codes) | set(data_sheet_codes))
    missing_from_input = sorted(set(target_codes) - set(clean_df["station_id"].unique().tolist()))
    if missing_from_input:
        readable = [f"{c}:{code_to_label.get(c, '')}" for c in missing_from_input]
        warnings.append("Stations absent in input and filled by --fill-missing: " + ", ".join(readable))

    matrix, output_stats = build_matrix(clean_df, target_codes, fill_missing)

    gaps = compute_time_gaps(clean_df["time"])
    if gaps:
        warnings.append(f"Detected non-hourly time gaps: {len(gaps)}")

    write_stats = write_output_template(wb_tpl, args.sheet_data, matrix, target_codes, args.data_start_row)
    wb_tpl.save(out_xlsx)
    logger.info("Output workbook written: %s", out_xlsx)

    for w in warnings:
        logger.warning(w)

    report = {
        "run_id": run_id,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "input_file": str(input_path.resolve()),
        "template_file": str(template_path.resolve()),
        "output_file": str(out_xlsx.resolve()),
        "log_file": str(log_path.resolve()),
        "input_stats": input_stats,
        "mapping_stats": {
            **mapping_stats,
            "template_station_count_stations_sheet": len(stations_sheet_codes),
            "template_station_count_data_sheet_before": len(data_sheet_codes),
            "target_station_count_after": len(target_codes),
            "stations_missing_from_input": missing_from_input,
        },
        "cleaning_stats": cleaning_stats,
        "time_quality": {
            "gap_count": len(gaps),
            "gaps": gaps,
        },
        "output_stats": {
            **output_stats,
            **write_stats,
        },
        "warnings": warnings,
        "strict_mode": bool(args.strict),
    }

    with out_json.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    with out_txt.open("w", encoding="utf-8") as f:
        f.write(report_txt(report))

    logger.info("Report JSON written: %s", out_json)
    logger.info("Report TXT written: %s", out_txt)

    if args.strict and warnings:
        logger.error("Strict mode enabled and warnings were detected.")
        return 1

    logger.info("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
