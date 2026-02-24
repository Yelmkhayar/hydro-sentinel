#!/usr/bin/env python3
"""
Prepare observed flow (m3/s) input into template_multi_station_flow_m3s.xlsx format.

Supports input:
- .xls exported as HTML table (common in this dataset)
- .xlsx DataTable-like sheets
- .csv (semicolon or comma)
"""

from __future__ import annotations

import argparse
import difflib
import json
import logging
import re
import unicodedata
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import openpyxl
import pandas as pd

TIMESTAMP_FMT = "%Y-%m-%dT%H:%M:%S"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("débit", "debit")
    s = s.replace("(m3/s)", "")
    s = s.replace("_", " ")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = " ".join(s.split())
    return s


def parse_fill_value(raw: str):
    low = raw.strip().lower()
    if low in {"", "nan", "none", "null"}:
        return None
    return float(raw)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Prepare observed flow data for BD ingestion template.")
    p.add_argument("--input", required=True, help="Input file (.xls/.xlsx/.csv).")
    p.add_argument("--template", required=True, help="Flow template path (.xlsx).")
    p.add_argument("--outdir", default="outputs/runs", help="Output directory.")
    p.add_argument("--sheet-data", default="Données", help="Template data sheet.")
    p.add_argument("--sheet-stations", default="Stations", help="Template stations sheet.")
    p.add_argument("--input-sheet", default="DataTable", help="Input sheet when --input is .xlsx.")
    p.add_argument("--data-start-row", type=int, default=4, help="First data row in template data sheet.")
    p.add_argument("--fill-missing", default="nan", help="Fill missing cells: nan|0|<float>.")
    p.add_argument("--resample-rule", default="1h", help="Resample rule (default: 1h).")
    p.add_argument(
        "--agg",
        default="mean",
        choices=["mean", "last", "sum", "max", "min", "median"],
        help="Aggregation for resampling.",
    )
    p.add_argument("--strict", action="store_true", help="Fail if warnings are present.")
    return p.parse_args()


def setup_logger(path: Path) -> logging.Logger:
    logger = logging.getLogger("prepare_flow_observed")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    sh = logging.StreamHandler()
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    fh = logging.FileHandler(path, encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger


def extract_template_stations(ws_st) -> Tuple[List[int], Dict[int, str], Dict[str, int]]:
    codes: List[int] = []
    code_to_name: Dict[int, str] = {}
    alias_to_code: Dict[str, int] = {}

    for r in range(2, ws_st.max_row + 1):
        code = ws_st.cell(r, 1).value
        name = ws_st.cell(r, 2).value
        if code is None or name is None:
            continue
        code_int = int(str(code).strip())
        name_str = str(name).strip()

        codes.append(code_int)
        code_to_name[code_int] = name_str

        aliases = {
            normalize_text(name_str),
            normalize_text(f"{name_str} debit"),
            normalize_text(f"{name_str} débit"),
            normalize_text(f"{name_str}_Debit (m3/s)"),
            normalize_text(f"{name_str}_Débit (m3/s)"),
        }
        for a in aliases:
            if a and a not in alias_to_code:
                alias_to_code[a] = code_int

    # Add explicit business aliases seen in observed files.
    # Key = observed label variant, Value = canonical station name in template.
    manual_alias_to_station = {
        "brg de garde debit": "Bge Garde de Sebou",
        "barrage de garde debit": "Bge Garde de Sebou",
        "pont elmalha debit": "El Malha",
        "pont el malha debit": "El Malha",
        "pont sebbou debit": "Ain Sebou",
    }
    name_to_code = {normalize_text(v): k for k, v in code_to_name.items()}
    for alias_label, target_station_name in manual_alias_to_station.items():
        target_code = name_to_code.get(normalize_text(target_station_name))
        alias_norm = normalize_text(alias_label)
        if target_code is not None and alias_norm and alias_norm not in alias_to_code:
            alias_to_code[alias_norm] = target_code

    return sorted(set(codes)), code_to_name, alias_to_code


def extract_template_data_codes(ws_data) -> List[int]:
    out: List[int] = []
    for c in range(2, ws_data.max_column + 1):
        v = ws_data.cell(3, c).value
        if isinstance(v, (int, float)):
            out.append(int(v))
    return out


def parse_html_xls(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8", errors="ignore")
    rows: List[List[str]] = []

    for tr in re.findall(r"<tr>(.*?)</tr>", text, flags=re.S | re.I):
        cells = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", tr, flags=re.S | re.I)
        cleaned = [unescape(re.sub(r"<[^>]+>", "", cell)).strip() for cell in cells]
        if cleaned:
            rows.append(cleaned)

    if len(rows) < 2:
        raise ValueError("Could not parse HTML table from .xls input.")

    header = rows[0]
    data = rows[1:]
    width = len(header)
    data = [r for r in data if len(r) == width]
    df = pd.DataFrame(data, columns=header)
    return df


def parse_xlsx(path: Path, sheet_name: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Input sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    # Support two layouts:
    # A) row1='Code', row2 labels, row3+ values
    # B) row1 headers, row2+ values
    first_cell = normalize_text(ws.cell(1, 1).value)

    rows: List[dict] = []
    if first_cell == "code":
        col_headers: List[Tuple[int, str]] = []
        for c in range(2, ws.max_column + 1):
            label = ws.cell(2, c).value
            col_headers.append((c, str(label) if label is not None else ""))

        for r in range(3, ws.max_row + 1):
            t = ws.cell(r, 1).value
            if t is None:
                continue
            for c, label in col_headers:
                rows.append({"time": t, "source_col": label, "value": ws.cell(r, c).value})
    else:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        dt_col = str(headers[0]) if headers and headers[0] is not None else "Date et heure"
        col_headers = [str(h) if h is not None else "" for h in headers[1:]]

        for r in range(2, ws.max_row + 1):
            t = ws.cell(r, 1).value
            if t is None:
                continue
            for idx, label in enumerate(col_headers, start=2):
                rows.append({"time": t, "source_col": label, "value": ws.cell(r, idx).value})

    return pd.DataFrame(rows)


def parse_csv(path: Path) -> pd.DataFrame:
    first_line = path.read_text(encoding="utf-8", errors="ignore").splitlines()[0]
    sep = ";" if first_line.count(";") >= first_line.count(",") else ","
    raw = pd.read_csv(path, sep=sep, dtype=str)
    if raw.shape[1] < 2:
        raise ValueError("CSV needs one time column + station columns.")

    time_col = raw.columns[0]
    station_cols = list(raw.columns[1:])
    melted = raw.melt(id_vars=[time_col], value_vars=station_cols, var_name="source_col", value_name="value")
    return melted.rename(columns={time_col: "time"})


def parse_input(path: Path, input_sheet: str) -> Tuple[pd.DataFrame, str]:
    ext = path.suffix.lower()
    if ext == ".xls":
        return parse_html_xls(path), "html_xls"
    if ext == ".xlsx":
        return parse_xlsx(path, input_sheet), "xlsx"
    if ext == ".csv":
        return parse_csv(path), "csv"
    raise ValueError(f"Unsupported input extension: {ext}")


def map_columns_to_codes(
    df: pd.DataFrame,
    alias_to_code: Dict[str, int],
    code_to_name: Dict[int, str],
    warnings: List[str],
) -> Tuple[pd.DataFrame, Dict[str, object]]:
    out = df.copy()

    source_cols = sorted(out["source_col"].dropna().astype(str).unique().tolist())
    known_aliases = list(alias_to_code.keys())

    col_to_code: Dict[str, int] = {}
    exact_matches = 0
    fuzzy_matches = 0
    fuzzy_details: List[str] = []
    unmapped_cols: List[str] = []

    for col in source_cols:
        norm = normalize_text(col)
        code = alias_to_code.get(norm)
        if code is not None:
            col_to_code[col] = code
            exact_matches += 1
            continue

        # fuzzy fallback to absorb spelling drift (example: Zerarda -> Zrarda)
        if known_aliases:
            best = difflib.get_close_matches(norm, known_aliases, n=1, cutoff=0.82)
            if best:
                b = best[0]
                code = alias_to_code[b]
                col_to_code[col] = code
                fuzzy_matches += 1
                fuzzy_details.append(f"{col} -> code {code} ({code_to_name.get(code,'')}) via '{b}'")
                continue

        unmapped_cols.append(col)

    if fuzzy_details:
        warnings.append("Fuzzy station mapping used: " + " | ".join(fuzzy_details))
    if unmapped_cols:
        warnings.append("Unmapped input station columns ignored: " + ", ".join(unmapped_cols))

    out["station_id"] = out["source_col"].map(col_to_code)
    dropped = int(out["station_id"].isna().sum())
    out = out.dropna(subset=["station_id"]).copy()
    out["station_id"] = out["station_id"].astype(int)

    stats = {
        "source_station_columns": len(source_cols),
        "exact_column_matches": exact_matches,
        "fuzzy_column_matches": fuzzy_matches,
        "unmapped_station_columns": unmapped_cols,
        "rows_dropped_unmapped": dropped,
    }
    return out, stats


def clean_and_resample(
    df: pd.DataFrame,
    resample_rule: str,
    agg: str,
    warnings: List[str],
) -> Tuple[pd.DataFrame, Dict[str, object]]:
    out = df.copy()

    out["time"] = pd.to_datetime(out["time"], errors="coerce", dayfirst=True)
    bad_time = int(out["time"].isna().sum())
    if bad_time:
        warnings.append(f"Invalid timestamps dropped: {bad_time}")
        out = out.dropna(subset=["time"])

    out["value"] = out["value"].astype(str).str.replace(",", ".", regex=False)
    out["value"] = pd.to_numeric(out["value"], errors="coerce")
    bad_val = int(out["value"].isna().sum())
    if bad_val:
        warnings.append(f"Non-numeric flow values converted to NaN: {bad_val}")

    negative_count = int((out["value"] < 0).sum(skipna=True))
    if negative_count:
        warnings.append(f"Negative flow values detected: {negative_count}")

    out = out.sort_values(["station_id", "time"]).reset_index(drop=True)

    # Detect native step before resampling
    native_steps = out[["station_id", "time"]].dropna()
    step_minutes = None
    if not native_steps.empty:
        d = (
            native_steps.groupby("station_id")["time"].diff().dropna().dt.total_seconds() / 60.0
        )
        if len(d) > 0:
            step_minutes = float(d.median())

    if resample_rule:
        grouped = out.set_index("time").groupby("station_id")["value"]
        if agg == "mean":
            rs = grouped.resample(resample_rule).mean()
        elif agg == "last":
            rs = grouped.resample(resample_rule).last()
        elif agg == "sum":
            rs = grouped.resample(resample_rule).sum(min_count=1)
        elif agg == "max":
            rs = grouped.resample(resample_rule).max()
        elif agg == "min":
            rs = grouped.resample(resample_rule).min()
        elif agg == "median":
            rs = grouped.resample(resample_rule).median()
        else:
            raise ValueError(f"Unsupported aggregation: {agg}")

        out = rs.reset_index().rename(columns={"value": "flow"})
    else:
        out = out.rename(columns={"value": "flow"})

    dup = out.duplicated(subset=["time", "station_id"], keep="last")
    dup_count = int(dup.sum())
    if dup_count:
        warnings.append(f"Duplicate (time, station_id) after resample dropped: {dup_count}")
        out = out.loc[~dup]

    out = out.sort_values(["time", "station_id"]).reset_index(drop=True)

    stats = {
        "rows_after_cleaning_and_resample": int(len(out)),
        "native_step_minutes_median": step_minutes,
        "resample_rule": resample_rule,
        "resample_agg": agg,
        "invalid_time_rows": bad_time,
        "invalid_value_rows": bad_val,
        "negative_value_rows": negative_count,
        "duplicate_rows_removed": dup_count,
        "time_min": out["time"].min().isoformat() if len(out) else None,
        "time_max": out["time"].max().isoformat() if len(out) else None,
        "station_count": int(out["station_id"].nunique()) if len(out) else 0,
        "time_count": int(out["time"].nunique()) if len(out) else 0,
    }
    return out, stats


def compute_time_gaps(times: pd.Series, expected_step: str = "1h") -> List[str]:
    if times.empty:
        return []
    uniq = pd.Series(sorted(times.unique()))
    diffs = uniq.diff().dropna()
    target = pd.Timedelta(expected_step)
    gaps = diffs[diffs != target]
    details: List[str] = []
    for idx, delta in gaps.items():
        details.append(f"gap={delta} between {uniq.iloc[idx-1]} and {uniq.iloc[idx]}")
    return details


def build_matrix(df: pd.DataFrame, target_codes: Sequence[int], fill_missing) -> Tuple[pd.DataFrame, Dict[str, object]]:
    piv = df.pivot_table(index="time", columns="station_id", values="flow", aggfunc="last").sort_index()
    matrix = piv.reindex(columns=target_codes)
    pre_fill_missing = int(matrix.isna().sum().sum())
    if fill_missing is not None:
        matrix = matrix.fillna(fill_missing)

    flat = pd.Series(matrix.to_numpy().ravel())
    valid = flat.dropna()

    stats = {
        "output_rows": int(matrix.shape[0]),
        "output_station_columns": int(matrix.shape[1]),
        "pre_fill_missing_cells": pre_fill_missing,
        "fill_missing_value": fill_missing,
        "flow_min": float(valid.min()) if len(valid) else None,
        "flow_max": float(valid.max()) if len(valid) else None,
        "flow_mean": float(valid.mean()) if len(valid) else None,
        "flow_p95": float(valid.quantile(0.95)) if len(valid) else None,
        "flow_p99": float(valid.quantile(0.99)) if len(valid) else None,
    }
    return matrix, stats


def clear_data_region(ws, start_row: int, max_col: int) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c, value=None)


def write_template(wb, sheet_data: str, matrix: pd.DataFrame, target_codes: Sequence[int], start_row: int) -> Dict[str, int]:
    ws = wb[sheet_data]
    ws.cell(3, 1, value="timestamp")
    for i, code in enumerate(target_codes, start=2):
        ws.cell(3, i, value=int(code))

    max_col = len(target_codes) + 1
    clear_data_region(ws, start_row, max_col)

    r = start_row
    for ts, row in matrix.iterrows():
        ws.cell(r, 1, value=pd.Timestamp(ts).strftime(TIMESTAMP_FMT))
        for c, v in enumerate(row.tolist(), start=2):
            ws.cell(r, c, value=None if pd.isna(v) else float(v))
        r += 1

    return {
        "written_rows": int(matrix.shape[0]),
        "written_station_columns": int(matrix.shape[1]),
        "first_output_row": int(start_row),
        "last_output_row": int(r - 1),
    }


def render_report_txt(report: Dict[str, object]) -> str:
    lines: List[str] = []
    lines.append("Transformation Report - Observed Flow (m3/s)")
    lines.append(f"run_id: {report['run_id']}")
    lines.append(f"input_file: {report['input_file']}")
    lines.append(f"input_format: {report['input_format']}")
    lines.append(f"template_file: {report['template_file']}")
    lines.append(f"output_file: {report['output_file']}")
    lines.append("")

    lines.append("Mapping Stats")
    for k, v in sorted(report["mapping_stats"].items()):
        lines.append(f"- {k}: {v}")

    lines.append("")
    lines.append("Processing Stats")
    for k, v in sorted(report["processing_stats"].items()):
        lines.append(f"- {k}: {v}")

    lines.append("")
    lines.append("Output Stats")
    for k, v in sorted(report["output_stats"].items()):
        lines.append(f"- {k}: {v}")

    lines.append("")
    lines.append("Time Quality")
    lines.append(f"- gap_count: {report['time_quality']['gap_count']}")
    for g in report["time_quality"]["gaps"]:
        lines.append(f"  * {g}")

    lines.append("")
    lines.append("Warnings")
    if report["warnings"]:
        for w in report["warnings"]:
            lines.append(f"- {w}")
    else:
        lines.append("- none")

    return "\n".join(lines) + "\n"


def main() -> int:
    args = parse_args()
    fill_missing = parse_fill_value(args.fill_missing)

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    input_path = Path(args.input)
    template_path = Path(args.template)
    if not input_path.exists():
        print(f"ERROR input not found: {input_path}")
        return 2
    if not template_path.exists():
        print(f"ERROR template not found: {template_path}")
        return 2

    base = f"flow_observed_{run_id}"
    out_xlsx = outdir / f"template_multi_station_flow_m3s_observed_{run_id}.xlsx"
    out_json = outdir / f"{base}_report.json"
    out_txt = outdir / f"{base}_report.txt"
    log_path = outdir / f"{base}.log"

    logger = setup_logger(log_path)
    warnings: List[str] = []

    logger.info("Loading template: %s", template_path)
    wb = openpyxl.load_workbook(template_path)
    if args.sheet_data not in wb.sheetnames:
        logger.error("Template data sheet not found: %s", args.sheet_data)
        return 1
    if args.sheet_stations not in wb.sheetnames:
        logger.error("Template stations sheet not found: %s", args.sheet_stations)
        return 1

    ws_data = wb[args.sheet_data]
    ws_st = wb[args.sheet_stations]

    station_codes_sheet, code_to_name, alias_to_code = extract_template_stations(ws_st)
    existing_data_codes = extract_template_data_codes(ws_data)

    logger.info("Reading input: %s", input_path)
    raw_df, input_format = parse_input(input_path, args.input_sheet)
    if raw_df.empty:
        logger.error("Input parsed with zero rows.")
        return 1

    # Harmonize expected columns
    if "source_col" not in raw_df.columns:
        cols = list(raw_df.columns)
        time_col = cols[0]
        melt = raw_df.melt(id_vars=[time_col], var_name="source_col", value_name="value")
        raw_df = melt.rename(columns={time_col: "time"})
    elif "value" not in raw_df.columns:
        # fallback if parser used original metric columns
        metric_cols = [c for c in raw_df.columns if c not in {"time", "source_col"}]
        if metric_cols:
            raw_df = raw_df.rename(columns={metric_cols[0]: "value"})

    mapped_df, mapping_stats = map_columns_to_codes(raw_df[["time", "source_col", "value"]], alias_to_code, code_to_name, warnings)
    processed_df, proc_stats = clean_and_resample(mapped_df, args.resample_rule, args.agg, warnings)

    if processed_df.empty:
        logger.error("No data left after mapping/cleaning/resampling.")
        return 1

    target_codes = sorted(set(station_codes_sheet) | set(existing_data_codes))

    missing_from_input = sorted(set(target_codes) - set(processed_df["station_id"].unique().tolist()))
    if missing_from_input:
        label_list = [f"{c}:{code_to_name.get(c,'') or 'unknown'}" for c in missing_from_input]
        warnings.append("Stations absent in input and filled by --fill-missing: " + ", ".join(label_list))

    matrix, matrix_stats = build_matrix(processed_df, target_codes, fill_missing)
    gaps = compute_time_gaps(processed_df["time"], expected_step=args.resample_rule)
    if gaps:
        warnings.append(f"Detected non-{args.resample_rule} time gaps: {len(gaps)}")

    write_stats = write_template(wb, args.sheet_data, matrix, target_codes, args.data_start_row)
    wb.save(out_xlsx)
    logger.info("Output workbook written: %s", out_xlsx)

    for w in warnings:
        logger.warning(w)

    report = {
        "run_id": run_id,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "input_file": str(input_path.resolve()),
        "input_format": input_format,
        "template_file": str(template_path.resolve()),
        "output_file": str(out_xlsx.resolve()),
        "log_file": str(log_path.resolve()),
        "mapping_stats": {
            **mapping_stats,
            "template_station_count_stations_sheet": len(station_codes_sheet),
            "template_station_count_data_sheet_before": len(existing_data_codes),
            "target_station_count_after": len(target_codes),
            "stations_missing_from_input": missing_from_input,
        },
        "processing_stats": proc_stats,
        "time_quality": {
            "gap_count": len(gaps),
            "gaps": gaps,
        },
        "output_stats": {
            **matrix_stats,
            **write_stats,
        },
        "warnings": warnings,
        "strict_mode": bool(args.strict),
    }

    with out_json.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    with out_txt.open("w", encoding="utf-8") as f:
        f.write(render_report_txt(report))

    logger.info("Report JSON written: %s", out_json)
    logger.info("Report TXT written: %s", out_txt)

    if args.strict and warnings:
        logger.error("Strict mode enabled and warnings were detected.")
        return 1

    logger.info("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
