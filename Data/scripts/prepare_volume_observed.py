#!/usr/bin/env python3
"""Prepare observed reservoir volume data into template_multi_station_volume_hm3.xlsx format."""

from __future__ import annotations

import argparse
import difflib
import json
import logging
import re
import unicodedata
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

import openpyxl
import pandas as pd

TIMESTAMP_FMT = "%Y-%m-%dT%H:%M:%S"


def normalize_text(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("_", " ")
    s = s.replace("(mm3)", "").replace("(mm 3)", "")
    s = s.replace("(hm3)", "").replace("(hm 3)", "")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return " ".join(s.split())


def parse_fill(raw: str):
    low = raw.strip().lower()
    if low in {"", "nan", "none", "null"}:
        return None
    return float(raw)


def setup_logger(path: Path) -> logging.Logger:
    logger = logging.getLogger("prepare_volume_observed")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    sh = logging.StreamHandler()
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    fh = logging.FileHandler(path, encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Prepare observed volume data for BD ingestion template.")
    p.add_argument("--input", required=True)
    p.add_argument("--template", required=True)
    p.add_argument("--outdir", default="outputs/runs")
    p.add_argument("--sheet-data", default="Données")
    p.add_argument("--sheet-stations", default="Stations")
    p.add_argument("--resample-rule", default="1h")
    p.add_argument("--agg", default="mean", choices=["mean", "last", "min", "max", "median"])
    p.add_argument("--fill-missing", default="nan")
    p.add_argument("--data-start-row", type=int, default=4)
    p.add_argument("--strict", action="store_true")
    return p.parse_args()


def extract_template_stations(ws_st) -> Tuple[List[int], Dict[int, str], Dict[str, int]]:
    codes: List[int] = []
    code_to_name: Dict[int, str] = {}
    alias_to_code: Dict[str, int] = {}
    for r in range(2, ws_st.max_row + 1):
        code = ws_st.cell(r, 1).value
        name = ws_st.cell(r, 2).value
        if code is None or name is None:
            continue
        code_i = int(str(code).strip())
        name_s = str(name).strip()
        codes.append(code_i)
        code_to_name[code_i] = name_s
        aliases = {
            normalize_text(name_s),
            normalize_text(f"{name_s} volume"),
            normalize_text(f"brg {name_s} volume"),
            normalize_text(f"{name_s}_volume"),
        }
        for a in aliases:
            if a and a not in alias_to_code:
                alias_to_code[a] = code_i
    return sorted(set(codes)), code_to_name, alias_to_code


def extract_data_codes(ws_data) -> List[int]:
    codes: List[int] = []
    for c in range(2, ws_data.max_column + 1):
        v = ws_data.cell(3, c).value
        if isinstance(v, (int, float)):
            codes.append(int(v))
    return codes


def parse_html_xls(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8", errors="ignore")
    rows = []
    for tr in re.findall(r"<tr>(.*?)</tr>", text, flags=re.S | re.I):
        cells = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", tr, flags=re.S | re.I)
        vals = [unescape(re.sub(r"<[^>]+>", "", c)).strip() for c in cells]
        if vals:
            rows.append(vals)
    if len(rows) < 2:
        raise ValueError("Cannot parse HTML table from input .xls")
    header = rows[0]
    width = len(header)
    body = [r for r in rows[1:] if len(r) == width]
    return pd.DataFrame(body, columns=header)


def parse_input(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext == ".xls":
        return parse_html_xls(path)
    if ext == ".csv":
        raw = pd.read_csv(path, sep=";", dtype=str)
        return raw
    if ext == ".xlsx":
        return pd.read_excel(path)
    raise ValueError(f"Unsupported input extension: {ext}")


def to_long_volume(df: pd.DataFrame, warnings: List[str]) -> Tuple[pd.DataFrame, Dict[str, object]]:
    cols = list(df.columns)
    if len(cols) < 2:
        raise ValueError("Input needs one datetime column and at least one metric column")

    time_col = cols[0]
    metric_cols = cols[1:]

    kept_volume_cols = []
    ignored_rate_cols = []
    ignored_other_cols = []

    for c in metric_cols:
        n = normalize_text(c)
        if "remplissage" in n or "taux" in n or "%" in str(c):
            ignored_rate_cols.append(str(c))
        elif "volume" in n:
            kept_volume_cols.append(str(c))
        else:
            ignored_other_cols.append(str(c))

    if ignored_rate_cols:
        warnings.append("Rate/fill columns ignored for now: " + ", ".join(ignored_rate_cols))
    if ignored_other_cols:
        warnings.append("Non-volume columns ignored: " + ", ".join(ignored_other_cols))
    if not kept_volume_cols:
        raise ValueError("No volume column detected in input")

    long_df = df[[time_col] + kept_volume_cols].copy()
    long_df = long_df.melt(id_vars=[time_col], var_name="source_col", value_name="value")
    long_df = long_df.rename(columns={time_col: "time"})

    stats = {
        "input_metric_columns": len(metric_cols),
        "volume_columns_kept": kept_volume_cols,
        "rate_columns_ignored": ignored_rate_cols,
        "other_columns_ignored": ignored_other_cols,
        "long_rows": int(len(long_df)),
    }
    return long_df, stats


def map_stations(long_df: pd.DataFrame, alias_to_code: Dict[str, int], code_to_name: Dict[int, str], warnings: List[str]) -> Tuple[pd.DataFrame, Dict[str, object]]:
    source_cols = sorted(long_df["source_col"].astype(str).unique().tolist())
    aliases = list(alias_to_code.keys())

    col_to_code: Dict[str, int] = {}
    exact = 0
    fuzzy = 0
    fuzzy_lines = []
    unmapped = []

    for c in source_cols:
        n = normalize_text(c)
        code = alias_to_code.get(n)
        if code is not None:
            col_to_code[c] = code
            exact += 1
            continue
        best = difflib.get_close_matches(n, aliases, n=1, cutoff=0.8)
        if best:
            b = best[0]
            code = alias_to_code[b]
            col_to_code[c] = code
            fuzzy += 1
            fuzzy_lines.append(f"{c} -> code {code} ({code_to_name.get(code,'')}) via '{b}'")
        else:
            unmapped.append(c)

    if fuzzy_lines:
        warnings.append("Fuzzy station mapping used: " + " | ".join(fuzzy_lines))
    if unmapped:
        warnings.append("Unmapped volume columns ignored: " + ", ".join(unmapped))

    out = long_df.copy()
    out["station_id"] = out["source_col"].map(col_to_code)
    dropped = int(out["station_id"].isna().sum())
    out = out.dropna(subset=["station_id"]).copy()
    out["station_id"] = out["station_id"].astype(int)

    stats = {
        "source_volume_columns": len(source_cols),
        "exact_column_matches": exact,
        "fuzzy_column_matches": fuzzy,
        "unmapped_volume_columns": unmapped,
        "rows_dropped_unmapped": dropped,
    }
    return out, stats


def clean_and_resample(df: pd.DataFrame, rule: str, agg: str, warnings: List[str]) -> Tuple[pd.DataFrame, Dict[str, object]]:
    out = df.copy()
    out["time"] = pd.to_datetime(out["time"], dayfirst=True, errors="coerce")
    bad_time = int(out["time"].isna().sum())
    if bad_time:
        warnings.append(f"Invalid timestamps dropped: {bad_time}")
        out = out.dropna(subset=["time"])

    out["value"] = out["value"].astype(str).str.replace(",", ".", regex=False)
    out["value"] = pd.to_numeric(out["value"], errors="coerce")
    bad_val = int(out["value"].isna().sum())
    if bad_val:
        warnings.append(f"Non-numeric volume values converted to NaN: {bad_val}")

    neg = int((out["value"] < 0).sum(skipna=True))
    if neg:
        warnings.append(f"Negative volume values detected: {neg}")

    out = out.sort_values(["station_id", "time"]).reset_index(drop=True)

    steps = out.groupby("station_id")["time"].diff().dropna().dt.total_seconds().div(60)
    step_median = float(steps.median()) if len(steps) else None

    grouped = out.set_index("time").groupby("station_id")["value"]
    if agg == "mean":
        rs = grouped.resample(rule).mean()
    elif agg == "last":
        rs = grouped.resample(rule).last()
    elif agg == "min":
        rs = grouped.resample(rule).min()
    elif agg == "max":
        rs = grouped.resample(rule).max()
    else:
        rs = grouped.resample(rule).median()
    out = rs.reset_index().rename(columns={"value": "volume"})

    stats = {
        "rows_after_cleaning_and_resample": int(len(out)),
        "native_step_minutes_median": step_median,
        "resample_rule": rule,
        "resample_agg": agg,
        "invalid_time_rows": bad_time,
        "invalid_value_rows": bad_val,
        "negative_value_rows": neg,
        "time_min": out["time"].min().isoformat() if len(out) else None,
        "time_max": out["time"].max().isoformat() if len(out) else None,
        "station_count": int(out["station_id"].nunique()) if len(out) else 0,
        "time_count": int(out["time"].nunique()) if len(out) else 0,
    }
    return out, stats


def build_matrix(df: pd.DataFrame, target_codes: Sequence[int], fill_missing) -> Tuple[pd.DataFrame, Dict[str, object]]:
    piv = df.pivot_table(index="time", columns="station_id", values="volume", aggfunc="last").sort_index()
    matrix = piv.reindex(columns=target_codes)
    pre_fill = int(matrix.isna().sum().sum())
    if fill_missing is not None:
        matrix = matrix.fillna(fill_missing)

    valid = pd.Series(matrix.to_numpy().ravel()).dropna()
    stats = {
        "output_rows": int(matrix.shape[0]),
        "output_station_columns": int(matrix.shape[1]),
        "pre_fill_missing_cells": pre_fill,
        "fill_missing_value": fill_missing,
        "volume_min": float(valid.min()) if len(valid) else None,
        "volume_max": float(valid.max()) if len(valid) else None,
        "volume_mean": float(valid.mean()) if len(valid) else None,
        "volume_p95": float(valid.quantile(0.95)) if len(valid) else None,
        "volume_p99": float(valid.quantile(0.99)) if len(valid) else None,
    }
    return matrix, stats


def time_gaps(times: pd.Series, expected: str) -> List[str]:
    if times.empty:
        return []
    uniq = pd.Series(sorted(times.unique()))
    diff = uniq.diff().dropna()
    target = pd.Timedelta(expected)
    gaps = diff[diff != target]
    out = []
    for i, d in gaps.items():
        out.append(f"gap={d} between {uniq.iloc[i-1]} and {uniq.iloc[i]}")
    return out


def clear_region(ws, start_row: int, max_col: int) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c, value=None)


def write_template(wb, sheet: str, matrix: pd.DataFrame, target_codes: Sequence[int], start_row: int) -> Dict[str, int]:
    ws = wb[sheet]
    ws.cell(3, 1, value="timestamp")
    for i, code in enumerate(target_codes, start=2):
        ws.cell(3, i, value=int(code))

    max_col = len(target_codes) + 1
    clear_region(ws, start_row, max_col)

    row = start_row
    for ts, vals in matrix.iterrows():
        ws.cell(row, 1, value=pd.Timestamp(ts).strftime(TIMESTAMP_FMT))
        for c, v in enumerate(vals.tolist(), start=2):
            ws.cell(row, c, value=None if pd.isna(v) else float(v))
        row += 1

    return {
        "written_rows": int(matrix.shape[0]),
        "written_station_columns": int(matrix.shape[1]),
        "first_output_row": int(start_row),
        "last_output_row": int(row - 1),
    }


def report_txt(rep: Dict[str, object]) -> str:
    lines = [
        "Transformation Report - Observed Volume (hm3)",
        f"run_id: {rep['run_id']}",
        f"input_file: {rep['input_file']}",
        f"template_file: {rep['template_file']}",
        f"output_file: {rep['output_file']}",
        "",
        "Input Parsing",
    ]
    for k, v in sorted(rep["input_parsing"].items()):
        lines.append(f"- {k}: {v}")
    lines += ["", "Mapping Stats"]
    for k, v in sorted(rep["mapping_stats"].items()):
        lines.append(f"- {k}: {v}")
    lines += ["", "Processing Stats"]
    for k, v in sorted(rep["processing_stats"].items()):
        lines.append(f"- {k}: {v}")
    lines += ["", "Output Stats"]
    for k, v in sorted(rep["output_stats"].items()):
        lines.append(f"- {k}: {v}")
    lines += ["", "Time Quality"]
    lines.append(f"- gap_count: {rep['time_quality']['gap_count']}")
    for g in rep["time_quality"]["gaps"]:
        lines.append(f"  * {g}")
    lines += ["", "Warnings"]
    if rep["warnings"]:
        for w in rep["warnings"]:
            lines.append(f"- {w}")
    else:
        lines.append("- none")
    return "\n".join(lines) + "\n"


def main() -> int:
    args = parse_args()
    fill_missing = parse_fill(args.fill_missing)

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    inp = Path(args.input)
    tpl = Path(args.template)
    if not inp.exists():
        print(f"ERROR input not found: {inp}")
        return 2
    if not tpl.exists():
        print(f"ERROR template not found: {tpl}")
        return 2

    base = f"volume_observed_{run_id}"
    out_xlsx = outdir / f"template_multi_station_volume_hm3_observed_{run_id}.xlsx"
    out_json = outdir / f"{base}_report.json"
    out_txt = outdir / f"{base}_report.txt"
    log_path = outdir / f"{base}.log"

    logger = setup_logger(log_path)
    warnings: List[str] = []

    logger.info("Loading template: %s", tpl)
    wb = openpyxl.load_workbook(tpl)
    if args.sheet_data not in wb.sheetnames or args.sheet_stations not in wb.sheetnames:
        logger.error("Template sheets missing")
        return 1

    ws_data = wb[args.sheet_data]
    ws_st = wb[args.sheet_stations]

    station_codes_sheet, code_to_name, alias_to_code = extract_template_stations(ws_st)
    data_codes_before = extract_data_codes(ws_data)

    logger.info("Reading input: %s", inp)
    raw = parse_input(inp)

    long_df, parse_stats = to_long_volume(raw, warnings)
    mapped_df, map_stats = map_stations(long_df, alias_to_code, code_to_name, warnings)
    proc_df, proc_stats = clean_and_resample(mapped_df, args.resample_rule, args.agg, warnings)

    if proc_df.empty:
        logger.error("No valid data after processing")
        return 1

    target_codes = sorted(set(station_codes_sheet) | set(data_codes_before))
    missing = sorted(set(target_codes) - set(proc_df["station_id"].unique().tolist()))
    if missing:
        warnings.append(
            "Stations absent in input and filled by --fill-missing: "
            + ", ".join(f"{c}:{code_to_name.get(c,'')}" for c in missing)
        )

    matrix, out_stats = build_matrix(proc_df, target_codes, fill_missing)
    gaps = time_gaps(proc_df["time"], args.resample_rule)
    if gaps:
        warnings.append(f"Detected non-{args.resample_rule} time gaps: {len(gaps)}")

    write_stats = write_template(wb, args.sheet_data, matrix, target_codes, args.data_start_row)
    wb.save(out_xlsx)
    logger.info("Output workbook written: %s", out_xlsx)

    for w in warnings:
        logger.warning(w)

    report = {
        "run_id": run_id,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "input_file": str(inp.resolve()),
        "template_file": str(tpl.resolve()),
        "output_file": str(out_xlsx.resolve()),
        "log_file": str(log_path.resolve()),
        "input_parsing": parse_stats,
        "mapping_stats": {
            **map_stats,
            "template_station_count_stations_sheet": len(station_codes_sheet),
            "template_station_count_data_sheet_before": len(data_codes_before),
            "target_station_count_after": len(target_codes),
            "stations_missing_from_input": missing,
        },
        "processing_stats": proc_stats,
        "time_quality": {"gap_count": len(gaps), "gaps": gaps},
        "output_stats": {**out_stats, **write_stats},
        "warnings": warnings,
        "strict_mode": bool(args.strict),
    }

    out_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    out_txt.write_text(report_txt(report), encoding="utf-8")

    logger.info("Report JSON written: %s", out_json)
    logger.info("Report TXT written: %s", out_txt)

    if args.strict and warnings:
        logger.error("Strict mode enabled and warnings detected")
        return 1

    logger.info("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
