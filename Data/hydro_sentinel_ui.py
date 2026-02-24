#!/usr/bin/env python3
"""Hydro Sentinel desktop UI for running and analyzing data preparation workflows."""

from __future__ import annotations

import difflib
import queue
import re
import subprocess
import sys
import threading
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import openpyxl
import pandas as pd


@dataclass(frozen=True)
class WorkflowConfig:
    key: str
    title: str
    script_rel: str
    template_rel: str
    default_input_rel: str
    patterns: tuple[str, ...]
    needs_model: bool = False
    supports_resample: bool = False


WORKFLOWS = (
    WorkflowConfig(
        key="precip_model",
        title="Precipitation modele",
        script_rel="scripts/prepare_precip_model.py",
        template_rel="templates/precip/template_precip_multi_station_mm.xlsx",
        default_input_rel="data_raw/model/precip/stations",
        patterns=("*.csv",),
        needs_model=True,
    ),
    WorkflowConfig(
        key="precip_observed",
        title="Precipitation observee",
        script_rel="scripts/prepare_precip_observed.py",
        template_rel="templates/precip/template_precip_multi_station_mm.xlsx",
        default_input_rel="data_raw/observed/precip",
        patterns=("*.xlsx", "*.csv"),
    ),
    WorkflowConfig(
        key="flow_observed",
        title="Debit observe",
        script_rel="scripts/prepare_flow_observed.py",
        template_rel="templates/flow/template_flow_multi_station_m3s.xlsx",
        default_input_rel="data_raw/observed/flow",
        patterns=("*.xls", "*.xlsx", "*.csv"),
        supports_resample=True,
    ),
    WorkflowConfig(
        key="volume_observed",
        title="Volume observe",
        script_rel="scripts/prepare_volume_observed.py",
        template_rel="templates/volume/template_volume_multi_station_hm3.xlsx",
        default_input_rel="data_raw/observed/volume",
        patterns=("*.xls", "*.xlsx", "*.csv"),
        supports_resample=True,
    ),
)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("débit", "debit")
    s = s.replace("remplissage", "remplissage")
    s = s.replace("(m3/s)", "")
    s = s.replace("(mm3)", "")
    s = s.replace("(hm3)", "")
    s = s.replace("_", " ")
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return " ".join(s.split())


def parse_html_xls_to_df(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8", errors="ignore")
    rows = []
    for tr in re.findall(r"<tr>(.*?)</tr>", text, flags=re.S | re.I):
        cells = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", tr, flags=re.S | re.I)
        vals = [unescape(re.sub(r"<[^>]+>", "", c)).strip() for c in cells]
        if vals:
            rows.append(vals)
    if len(rows) < 2:
        return pd.DataFrame()
    header = rows[0]
    width = len(header)
    data = [r for r in rows[1:] if len(r) == width]
    return pd.DataFrame(data, columns=header)


class WorkflowTab:
    def __init__(self, app: "HydroSentinelUI", parent: ttk.Notebook, cfg: WorkflowConfig) -> None:
        self.app = app
        self.cfg = cfg
        self.frame = ttk.Frame(parent)
        parent.add(self.frame, text=cfg.title)

        self.mode_var = tk.StringVar(value="single")
        self.input_var = tk.StringVar(value=str(app.root_dir / cfg.default_input_rel))
        self.template_var = tk.StringVar(value=str(app.root_dir / cfg.template_rel))
        self.output_var = tk.StringVar(value=str(app.root_dir / "outputs/runs"))
        self.model_var = tk.StringVar(value="AROME")
        self.fill_var = tk.StringVar(value="nan")
        self.resample_var = tk.StringVar(value="1h")
        self.agg_var = tk.StringVar(value="mean")
        self.strict_var = tk.BooleanVar(value=False)
        self.selected_file_var = tk.StringVar(value="")

        self.detected_files: list[Path] = []
        self._build()

    def _build(self) -> None:
        frm = self.frame
        frm.columnconfigure(1, weight=1)

        row = 0
        ttk.Label(frm, text="Mode").grid(row=row, column=0, sticky="w", padx=8, pady=6)
        mfrm = ttk.Frame(frm)
        mfrm.grid(row=row, column=1, sticky="w", padx=8, pady=6)
        ttk.Radiobutton(mfrm, text="Single", value="single", variable=self.mode_var).pack(side="left", padx=6)
        ttk.Radiobutton(mfrm, text="Batch", value="batch", variable=self.mode_var).pack(side="left", padx=6)

        row += 1
        ttk.Label(frm, text="Input").grid(row=row, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(frm, textvariable=self.input_var).grid(row=row, column=1, sticky="ew", padx=8, pady=6)
        b1 = ttk.Frame(frm)
        b1.grid(row=row, column=2, sticky="w", padx=8, pady=6)
        ttk.Button(b1, text="Fichier", command=self._pick_input_file).pack(side="left", padx=2)
        ttk.Button(b1, text="Dossier", command=self._pick_input_dir).pack(side="left", padx=2)

        row += 1
        ttk.Label(frm, text="Template").grid(row=row, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(frm, textvariable=self.template_var).grid(row=row, column=1, sticky="ew", padx=8, pady=6)
        ttk.Button(frm, text="Parcourir", command=self._pick_template).grid(row=row, column=2, sticky="w", padx=8, pady=6)

        row += 1
        ttk.Label(frm, text="Output").grid(row=row, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(frm, textvariable=self.output_var).grid(row=row, column=1, sticky="ew", padx=8, pady=6)
        ttk.Button(frm, text="Parcourir", command=self._pick_output).grid(row=row, column=2, sticky="w", padx=8, pady=6)

        row += 1
        opts = ttk.LabelFrame(frm, text="Options")
        opts.grid(row=row, column=0, columnspan=3, sticky="ew", padx=8, pady=8)

        opt_row = 0
        if self.cfg.needs_model:
            ttk.Label(opts, text="Model").grid(row=opt_row, column=0, sticky="w", padx=6, pady=4)
            ttk.Entry(opts, textvariable=self.model_var, width=12).grid(row=opt_row, column=1, sticky="w", padx=6, pady=4)
            opt_row += 1

        if self.cfg.supports_resample:
            ttk.Label(opts, text="Resample").grid(row=opt_row, column=0, sticky="w", padx=6, pady=4)
            ttk.Entry(opts, textvariable=self.resample_var, width=12).grid(row=opt_row, column=1, sticky="w", padx=6, pady=4)
            ttk.Label(opts, text="Agg").grid(row=opt_row, column=2, sticky="w", padx=6, pady=4)
            ttk.Combobox(
                opts,
                textvariable=self.agg_var,
                state="readonly",
                width=10,
                values=("mean", "last", "min", "max", "median", "sum"),
            ).grid(row=opt_row, column=3, sticky="w", padx=6, pady=4)
            opt_row += 1

        ttk.Label(opts, text="Fill missing").grid(row=opt_row, column=0, sticky="w", padx=6, pady=4)
        ttk.Entry(opts, textvariable=self.fill_var, width=12).grid(row=opt_row, column=1, sticky="w", padx=6, pady=4)
        ttk.Checkbutton(opts, text="Strict", variable=self.strict_var).grid(row=opt_row, column=2, sticky="w", padx=6, pady=4)

        row += 1
        act = ttk.Frame(frm)
        act.grid(row=row, column=0, columnspan=3, sticky="ew", padx=8, pady=6)
        ttk.Button(act, text="Analyser Input", command=self.analyze_input).pack(side="left", padx=4)
        ttk.Button(act, text="Executer", command=self.run).pack(side="left", padx=4)

        row += 1
        ttk.Label(frm, text="Fichier (single si input=dossier)").grid(row=row, column=0, sticky="w", padx=8, pady=6)
        self.file_combo = ttk.Combobox(frm, textvariable=self.selected_file_var, state="readonly")
        self.file_combo.grid(row=row, column=1, sticky="ew", padx=8, pady=6)

        row += 1
        ttk.Label(frm, text="Patterns").grid(row=row, column=0, sticky="w", padx=8, pady=6)
        ttk.Label(frm, text=", ".join(self.cfg.patterns)).grid(row=row, column=1, sticky="w", padx=8, pady=6)

        row += 1
        analysis = ttk.LabelFrame(frm, text="Analyse input / Mapping")
        analysis.grid(row=row, column=0, columnspan=3, sticky="nsew", padx=8, pady=8)
        analysis.columnconfigure(0, weight=1)
        analysis.columnconfigure(1, weight=1)
        analysis.rowconfigure(1, weight=1)

        ttk.Label(analysis, text="Chiffres cles").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        ttk.Label(analysis, text="Mapping stations").grid(row=0, column=1, sticky="w", padx=6, pady=4)

        self.analysis_text = tk.Text(analysis, height=12, wrap="word")
        self.analysis_text.grid(row=1, column=0, sticky="nsew", padx=(6, 3), pady=6)

        map_frame = ttk.Frame(analysis)
        map_frame.grid(row=1, column=1, sticky="nsew", padx=(3, 6), pady=6)
        map_frame.columnconfigure(0, weight=1)
        map_frame.rowconfigure(0, weight=1)

        cols = ("source", "code", "station", "method")
        self.mapping_tree = ttk.Treeview(map_frame, columns=cols, show="headings", height=12)
        for c, w in (("source", 220), ("code", 60), ("station", 170), ("method", 90)):
            self.mapping_tree.heading(c, text=c)
            self.mapping_tree.column(c, width=w, anchor="w")
        yscroll = ttk.Scrollbar(map_frame, orient="vertical", command=self.mapping_tree.yview)
        self.mapping_tree.configure(yscrollcommand=yscroll.set)
        self.mapping_tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")

        frm.rowconfigure(row, weight=1)

    def _pick_input_file(self) -> None:
        p = filedialog.askopenfilename(initialdir=str(self.app.root_dir))
        if p:
            self.input_var.set(p)

    def _pick_input_dir(self) -> None:
        p = filedialog.askdirectory(initialdir=str(self.app.root_dir))
        if p:
            self.input_var.set(p)

    def _pick_template(self) -> None:
        p = filedialog.askopenfilename(initialdir=str(self.app.root_dir), filetypes=[("Excel", "*.xlsx")])
        if p:
            self.template_var.set(p)

    def _pick_output(self) -> None:
        p = filedialog.askdirectory(initialdir=str(self.app.root_dir))
        if p:
            self.output_var.set(p)

    def _detect_files(self) -> list[Path]:
        p = Path(self.input_var.get().strip())
        if not p.exists():
            return []
        if p.is_file():
            return [p]
        found: list[Path] = []
        for pat in self.cfg.patterns:
            found.extend(sorted(x for x in p.glob(pat) if x.is_file()))
        uniq: list[Path] = []
        seen = set()
        for f in found:
            key = str(f.resolve())
            if key not in seen:
                seen.add(key)
                uniq.append(f)
        return uniq

    def _load_template_mapping(self) -> tuple[dict[int, str], dict[str, int]]:
        code_to_name: dict[int, str] = {}
        alias_to_code: dict[str, int] = {}
        tpl = Path(self.template_var.get().strip())
        wb = openpyxl.load_workbook(tpl, data_only=True)
        ws = wb["Stations"] if "Stations" in wb.sheetnames else wb[wb.sheetnames[0]]

        for r in range(2, ws.max_row + 1):
            c = ws.cell(r, 1).value
            n = ws.cell(r, 2).value
            if c is None or n is None:
                continue
            code = int(str(c).strip())
            name = str(n).strip()
            code_to_name[code] = name
            aliases = {
                normalize_text(name),
                normalize_text(f"{name} debit"),
                normalize_text(f"{name} volume"),
                normalize_text(f"{name} pluie"),
                normalize_text(f"{name}_debit"),
                normalize_text(f"{name}_volume"),
                normalize_text(f"{name}_pluie"),
            }
            for a in aliases:
                if a and a not in alias_to_code:
                    alias_to_code[a] = code

        # business aliases seen in files
        manual = {
            "brg de garde debit": "Bge Garde de Sebou",
            "pont elmalha debit": "El Malha",
            "pont el malha debit": "El Malha",
            "pont sebbou debit": "Ain Sebou",
            "zerarda debit": "Zrarda",
        }
        name_to_code = {normalize_text(v): k for k, v in code_to_name.items()}
        for a, n in manual.items():
            k = name_to_code.get(normalize_text(n))
            if k is not None:
                alias_to_code[normalize_text(a)] = k

        return code_to_name, alias_to_code

    def _wide_from_file(self, path: Path) -> pd.DataFrame:
        ext = path.suffix.lower()
        if ext == ".xls":
            return parse_html_xls_to_df(path)
        if ext == ".xlsx":
            return pd.read_excel(path)
        if ext == ".csv":
            txt = path.read_text(encoding="utf-8", errors="ignore").splitlines()[0]
            sep = ";" if txt.count(";") >= txt.count(",") else ","
            return pd.read_csv(path, sep=sep, dtype=str)
        return pd.DataFrame()

    def _analyze_file(self, path: Path, code_to_name: dict[int, str], alias_to_code: dict[str, int]) -> tuple[dict, list[tuple[str, str, str, str]]]:
        info = {"file": path.name, "rows": 0, "dates": (None, None), "values": 0, "nan": 0, "neg": 0, "min": None, "max": None, "mean": None, "stations": 0, "step": None}
        mapping_rows: list[tuple[str, str, str, str]] = []

        if self.cfg.needs_model:
            df = pd.read_csv(path, sep=";")
            if df.empty:
                return info, mapping_rows
            df["time"] = pd.to_datetime(df["time"], errors="coerce")
            df["rr"] = pd.to_numeric(df["rr"], errors="coerce")
            info["rows"] = int(len(df))
            info["dates"] = (str(df["time"].min()), str(df["time"].max()))
            info["values"] = int(df["rr"].notna().sum())
            info["nan"] = int(df["rr"].isna().sum())
            info["neg"] = int((df["rr"] < 0).sum(skipna=True))
            info["min"] = float(df["rr"].min()) if df["rr"].notna().any() else None
            info["max"] = float(df["rr"].max()) if df["rr"].notna().any() else None
            info["mean"] = float(df["rr"].mean()) if df["rr"].notna().any() else None
            info["stations"] = int(df["station_id"].nunique()) if "station_id" in df.columns else 0
            ts = pd.Series(sorted(df["time"].dropna().unique()))
            if len(ts) > 1:
                info["step"] = float(ts.diff().dropna().dt.total_seconds().div(60).median())
            return info, mapping_rows

        wide = self._wide_from_file(path)
        if wide.empty or wide.shape[1] < 2:
            return info, mapping_rows

        time_col = wide.columns[0]
        metrics = list(wide.columns[1:])

        if self.cfg.key == "volume_observed":
            keep = []
            for c in metrics:
                n = normalize_text(c)
                if "volume" in n:
                    keep.append(c)
            metrics = keep
            if not metrics:
                return info, mapping_rows

        m = wide[[time_col] + metrics].copy()
        long = m.melt(id_vars=[time_col], value_vars=metrics, var_name="source_col", value_name="value")
        long = long.rename(columns={time_col: "time"})
        long["time"] = pd.to_datetime(long["time"], errors="coerce", dayfirst=True)
        long["value"] = long["value"].astype(str).str.replace(",", ".", regex=False)
        long["value"] = pd.to_numeric(long["value"], errors="coerce")

        uniq_cols = sorted(long["source_col"].dropna().astype(str).unique().tolist())
        known_aliases = list(alias_to_code.keys())
        mapped_codes = set()

        for col in uniq_cols:
            norm = normalize_text(col)
            code = alias_to_code.get(norm)
            method = "exact"
            if code is None:
                best = difflib.get_close_matches(norm, known_aliases, n=1, cutoff=0.82)
                if best:
                    code = alias_to_code[best[0]]
                    method = "fuzzy"
                else:
                    method = "unmapped"

            if code is None:
                mapping_rows.append((col, "", "", method))
            else:
                mapped_codes.add(code)
                mapping_rows.append((col, str(code), code_to_name.get(code, ""), method))

        info["rows"] = int(len(long))
        info["dates"] = (str(long["time"].min()), str(long["time"].max()))
        info["values"] = int(long["value"].notna().sum())
        info["nan"] = int(long["value"].isna().sum())
        info["neg"] = int((long["value"] < 0).sum(skipna=True))
        valid = long["value"].dropna()
        info["min"] = float(valid.min()) if len(valid) else None
        info["max"] = float(valid.max()) if len(valid) else None
        info["mean"] = float(valid.mean()) if len(valid) else None
        info["stations"] = len(mapped_codes)

        ts = pd.Series(sorted(long["time"].dropna().unique()))
        if len(ts) > 1:
            info["step"] = float(ts.diff().dropna().dt.total_seconds().div(60).median())

        return info, mapping_rows

    def analyze_input(self) -> None:
        self.detected_files = self._detect_files()
        names = [p.name for p in self.detected_files]
        self.file_combo["values"] = names
        if names and (self.selected_file_var.get() not in names):
            self.selected_file_var.set(names[0])
        if not names:
            self.selected_file_var.set("")

        self.app.log(f"[{self.cfg.title}] Analyse input")
        self.app.log(f"  Path: {self.input_var.get()}")
        self.app.log(f"  Detectes: {len(self.detected_files)}")

        self.analysis_text.delete("1.0", "end")
        for item in self.mapping_tree.get_children():
            self.mapping_tree.delete(item)

        if not self.detected_files:
            self.analysis_text.insert("end", "Aucun fichier detecte.\n")
            return

        code_to_name, alias_to_code = self._load_template_mapping()

        targets = self.detected_files
        if self.mode_var.get() == "single" and len(self.detected_files) > 1:
            chosen = self.selected_file_var.get().strip()
            if chosen:
                targets = [p for p in self.detected_files if p.name == chosen]
            else:
                targets = [self.detected_files[0]]

        summaries = []
        mapping_preview = []
        for idx, f in enumerate(targets):
            s, m = self._analyze_file(f, code_to_name, alias_to_code)
            summaries.append(s)
            if idx == 0:
                mapping_preview = m

        total_values = sum(x["values"] for x in summaries)
        total_nan = sum(x["nan"] for x in summaries)
        total_neg = sum(x["neg"] for x in summaries)
        stations = sorted({x["stations"] for x in summaries})

        lines = [
            f"Workflow: {self.cfg.title}",
            f"Fichiers analyses: {len(summaries)}",
            f"Valeurs numeriques: {total_values}",
            f"Valeurs NaN: {total_nan}",
            f"Valeurs negatives: {total_neg}",
            f"Stations identifiees (par fichier): {stations}",
            "",
        ]

        for s in summaries[:8]:
            lines.extend(
                [
                    f"- {s['file']}",
                    f"  rows={s['rows']} values={s['values']} nan={s['nan']} neg={s['neg']}",
                    f"  dates: {s['dates'][0]} -> {s['dates'][1]}",
                    f"  stations={s['stations']} step_min={s['step']}",
                    f"  min={s['min']} max={s['max']} mean={s['mean']}",
                    "",
                ]
            )
        if len(summaries) > 8:
            lines.append(f"... {len(summaries) - 8} fichiers supplementaires non affiches")

        self.analysis_text.insert("end", "\n".join(lines))

        for row in mapping_preview:
            self.mapping_tree.insert("", "end", values=row)

    def _validate(self) -> bool:
        script = self.app.root_dir / self.cfg.script_rel
        template = Path(self.template_var.get().strip())
        output = Path(self.output_var.get().strip())

        if not script.exists():
            messagebox.showerror("Erreur", f"Script introuvable:\n{script}")
            return False
        if not template.exists():
            messagebox.showerror("Erreur", f"Template introuvable:\n{template}")
            return False

        files = self._detect_files()
        if not files:
            messagebox.showerror("Erreur", "Aucun fichier input detecte.")
            return False
        output.mkdir(parents=True, exist_ok=True)

        if self.mode_var.get() == "single" and len(files) > 1 and not self.selected_file_var.get().strip():
            messagebox.showerror("Erreur", "Selectionne un fichier dans la liste (single mode).")
            return False

        return True

    def _build_cmd(self, input_file: Path) -> list[str]:
        cmd = [
            sys.executable,
            str((self.app.root_dir / self.cfg.script_rel).resolve()),
            "--input",
            str(input_file),
            "--template",
            self.template_var.get().strip(),
            "--outdir",
            self.output_var.get().strip(),
        ]

        if self.cfg.needs_model:
            model = self.model_var.get().strip() or "AROME"
            cmd += ["--model", model]

        fill = self.fill_var.get().strip()
        if fill:
            cmd += ["--fill-missing", fill]

        if self.cfg.supports_resample:
            rs = self.resample_var.get().strip() or "1h"
            agg = self.agg_var.get().strip() or "mean"
            cmd += ["--resample-rule", rs, "--agg", agg]

        if self.strict_var.get():
            cmd += ["--strict"]

        return cmd

    def run(self) -> None:
        if not self._validate():
            return

        files = self._detect_files()
        if self.mode_var.get() == "single":
            chosen = self.selected_file_var.get().strip()
            if chosen:
                pick = [f for f in files if f.name == chosen]
                files = pick if pick else [files[0]]
            else:
                files = [files[0]]

        consolidate = self.mode_var.get() == "batch" and self.cfg.key in {"flow_observed", "volume_observed"}
        self.app.run_commands(
            self.cfg.title,
            [self._build_cmd(f) for f in files],
            consolidate=consolidate,
            workflow_key=self.cfg.key,
            template_path=Path(self.template_var.get().strip()),
            outdir=Path(self.output_var.get().strip()),
        )


class HydroSentinelUI:
    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title("Hydro Sentinel - Workflow UI")
        self.root.geometry("1280x860")
        self.root_dir = Path(__file__).resolve().parent
        self.log_queue: queue.Queue[str] = queue.Queue()
        self.running = False

        self._build()
        self.root.after(120, self._flush_logs)

    def _build(self) -> None:
        top = ttk.Frame(self.root)
        top.pack(fill="x", padx=8, pady=8)
        ttk.Label(top, text="Hydro Sentinel Data Processing", font=("Segoe UI", 14, "bold")).pack(side="left")
        ttk.Label(top, text=f"Root: {self.root_dir}").pack(side="right")

        body = ttk.PanedWindow(self.root, orient="vertical")
        body.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        nb_frame = ttk.Frame(body)
        self.notebook = ttk.Notebook(nb_frame)
        self.notebook.pack(fill="both", expand=True)
        self.tabs = [WorkflowTab(self, self.notebook, cfg) for cfg in WORKFLOWS]

        log_frame = ttk.LabelFrame(body, text="Logs")
        self.log_text = tk.Text(log_frame, height=16, wrap="word")
        self.log_text.pack(fill="both", expand=True, padx=6, pady=6)

        body.add(nb_frame, weight=4)
        body.add(log_frame, weight=2)

    def log(self, msg: str) -> None:
        self.log_queue.put(msg)

    def _flush_logs(self) -> None:
        try:
            while True:
                line = self.log_queue.get_nowait()
                self.log_text.insert("end", line + "\n")
                self.log_text.see("end")
        except queue.Empty:
            pass
        self.root.after(120, self._flush_logs)

    def run_commands(
        self,
        title: str,
        commands: list[list[str]],
        consolidate: bool = False,
        workflow_key: str = "",
        template_path: Path | None = None,
        outdir: Path | None = None,
    ) -> None:
        if self.running:
            messagebox.showwarning("Execution", "Un traitement est deja en cours.")
            return
        if not commands:
            messagebox.showwarning("Execution", "Aucune commande a executer.")
            return

        confirm = messagebox.askyesno(
            "Confirmation",
            (
                f"Workflow: {title}\n"
                f"Fichiers a traiter: {len(commands)}\n"
                f"Mode consolidation: {'ON' if consolidate else 'OFF'}\n\n"
                "Lancer l'execution ?"
            ),
        )
        if not confirm:
            return

        self.running = True
        threading.Thread(
            target=self._run_worker,
            args=(title, commands, consolidate, workflow_key, template_path, outdir),
            daemon=True,
        ).start()

    def _run_worker(
        self,
        title: str,
        commands: list[list[str]],
        consolidate: bool,
        workflow_key: str,
        template_path: Path | None,
        outdir: Path | None,
    ) -> None:
        ok = 0
        ko = 0
        generated_xlsx: list[Path] = []
        self.log(f"=== START {title} | jobs={len(commands)} ===")

        for i, cmd in enumerate(commands, start=1):
            self.log(f"[{i}/{len(commands)}] RUN: {' '.join(cmd)}")
            existing_outputs: set[str] = set()
            pattern = self._output_pattern_for_workflow(workflow_key)
            if outdir is not None and pattern:
                existing_outputs = {str(p.resolve()) for p in outdir.glob(pattern)}
            try:
                proc = subprocess.Popen(
                    cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                )
                assert proc.stdout is not None
                for line in proc.stdout:
                    self.log(line.rstrip())
                rc = proc.wait()
            except Exception as exc:
                rc = 1
                self.log(f"ERROR launching process: {exc}")

            if rc == 0:
                ok += 1
                self.log(f"[{i}] DONE rc=0")
                if outdir is not None and pattern:
                    after_outputs = {str(p.resolve()): p for p in outdir.glob(pattern)}
                    new_outputs = [p for k, p in after_outputs.items() if k not in existing_outputs]
                    if new_outputs:
                        latest = max(new_outputs, key=lambda p: p.stat().st_mtime)
                        generated_xlsx.append(latest)
                        self.log(f"[{i}] Output detecte: {latest.name}")
            else:
                ko += 1
                self.log(f"[{i}] FAIL rc={rc}")

        consolidated_path: Path | None = None
        if consolidate and ok > 0 and template_path is not None and outdir is not None and generated_xlsx:
            try:
                consolidated_path = self._consolidate_outputs(
                    workflow_key=workflow_key,
                    template_path=template_path,
                    outdir=outdir,
                    generated_files=generated_xlsx,
                )
                self.log(f"[CONSOLIDATION] Fichier consolide: {consolidated_path}")
            except Exception as exc:
                ko += 1
                self.log(f"[CONSOLIDATION] ERROR: {exc}")

        self.log(f"=== END {title} | success={ok} failed={ko} ===")
        self.running = False
        extra = f"\nConsolide: {consolidated_path}" if consolidated_path else ""
        messagebox.showinfo("Execution terminee", f"{title}\nSuccess: {ok}\nFailed: {ko}{extra}")

    def _output_pattern_for_workflow(self, workflow_key: str) -> str:
        if workflow_key == "flow_observed":
            return "template_multi_station_flow_m3s_observed_*.xlsx"
        if workflow_key == "volume_observed":
            return "template_multi_station_volume_hm3_observed_*.xlsx"
        if workflow_key == "precip_observed":
            return "template_multi_station_precip_mm_observed_*.xlsx"
        if workflow_key == "precip_model":
            return "template_multi_station_precip_mm_*.xlsx"
        return ""

    def _consolidated_name(self, workflow_key: str) -> str:
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        if workflow_key == "flow_observed":
            return f"template_multi_station_flow_m3s_observed_consolidated_{ts}.xlsx"
        if workflow_key == "volume_observed":
            return f"template_multi_station_volume_hm3_observed_consolidated_{ts}.xlsx"
        return f"template_consolidated_{ts}.xlsx"

    def _consolidate_outputs(
        self,
        workflow_key: str,
        template_path: Path,
        outdir: Path,
        generated_files: list[Path],
    ) -> Path:
        wb_tpl = openpyxl.load_workbook(template_path)
        ws_tpl = wb_tpl["Données"]

        # Keep template order by header row.
        ordered_codes: list[int] = []
        for c in range(2, ws_tpl.max_column + 1):
            v = ws_tpl.cell(3, c).value
            if isinstance(v, (int, float)):
                ordered_codes.append(int(v))

        values_by_ts_code: dict[tuple[pd.Timestamp, int], float] = {}
        all_ts: set[pd.Timestamp] = set()

        for fp in generated_files:
            wb = openpyxl.load_workbook(fp, data_only=True)
            if "Données" not in wb.sheetnames:
                continue
            ws = wb["Données"]

            code_to_col: dict[int, int] = {}
            for c in range(2, ws.max_column + 1):
                v = ws.cell(3, c).value
                if isinstance(v, (int, float)):
                    code_to_col[int(v)] = c

            for r in range(4, ws.max_row + 1):
                tval = ws.cell(r, 1).value
                if tval in (None, ""):
                    continue
                ts = pd.to_datetime(tval, errors="coerce")
                if pd.isna(ts):
                    continue
                ts = pd.Timestamp(ts)
                all_ts.add(ts)

                for code in ordered_codes:
                    col = code_to_col.get(code)
                    if col is None:
                        continue
                    val = ws.cell(r, col).value
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        continue
                    values_by_ts_code[(ts, code)] = float(val)

        sorted_ts = sorted(all_ts)

        # Clear data region.
        max_col = len(ordered_codes) + 1
        for r in range(4, ws_tpl.max_row + 1):
            for c in range(1, max_col + 1):
                ws_tpl.cell(r, c, value=None)

        # Ensure header.
        ws_tpl.cell(3, 1, value="timestamp")
        for i, code in enumerate(ordered_codes, start=2):
            ws_tpl.cell(3, i, value=code)

        row = 4
        for ts in sorted_ts:
            ws_tpl.cell(row, 1, value=ts.strftime("%Y-%m-%dT%H:%M:%S"))
            for i, code in enumerate(ordered_codes, start=2):
                v = values_by_ts_code.get((ts, code))
                ws_tpl.cell(row, i, value=v if v is not None else None)
            row += 1

        out_path = outdir / self._consolidated_name(workflow_key)
        wb_tpl.save(out_path)
        return out_path

    def run(self) -> None:
        self.root.mainloop()


def main() -> int:
    HydroSentinelUI().run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
